#works with Python 3.9

from openpyxl import Workbook
import os
from openpyxl import load_workbook
from tkinter import filedialog as fd

filename = fd.askopenfilename()
print("File chosen: " + filename)
wb = load_workbook(filename, data_only=True)
round_number = int(input("Number of rounds want to extract: "))

sheetlist = wb.sheetnames
for asheet in sheetlist:
    curentSheet = wb[asheet]
    FolderSheetName = curentSheet.title
    if not os.path.exists(FolderSheetName):
        os.makedirs(FolderSheetName)
    datarows = len(curentSheet['A']) -1
    for i in range(round_number):
        FolderRoundName = curentSheet.cell(row=1, column=i*6+2).value.split()[1] + 'ul'
        if not os.path.exists(FolderSheetName +'/'+ FolderRoundName):
            os.makedirs(FolderSheetName +'/'+ FolderRoundName)
        with open(FolderSheetName +'/'+ FolderRoundName + '/' + 'data.txt', 'w') as f:
            f.write(str(datarows) + "\n")
            for textrow in range(datarows):
                content = str(curentSheet.cell(row=textrow+2, column=i*6+6).value) + "   " + str(curentSheet.cell(row=textrow+2, column=i*6+7).value) + "   " + str(curentSheet.cell(row=textrow+2, column=1).value)
                f.write(content + "\n")
        with open(FolderSheetName +'/'+ FolderRoundName + '/' + 'fit.txt', 'w') as f:
            f.write("")
        with open(FolderSheetName +'/'+ FolderRoundName + '/' + 'param.txt', 'w') as f:
            f.write("")
        #print(curentSheet.cell(row=1, column=textrow*6+2).value)